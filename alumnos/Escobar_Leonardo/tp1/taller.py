import os
from datetime import datetime
from openpyxl import Workbook

# ===== FUNCIONES =====

def leer_numero(texto):
    valor = input(texto)
    valor = valor.replace(".", "")
    valor = valor.replace(",", ".")
    return float(valor)

def formatear(n):
    return f"${int(n):,}".replace(",", ".")

# ===== VARIABLES =====

trabajos = []
porc_empleado = 0.20
porc_inversion = 0.20
porc_gastos = 0.20

# ===== MENÚ =====

while True:
    print("\n===== SISTEMA TALLER PRO MAX =====")
    print("1. Cargar trabajo")
    print("2. Ver resumen")
    print("3. Cambiar porcentajes")
    print("4. Objetivo de ganancia")
    print("5. Exportar a Excel PRO")
    print("6. Salir")

    op = input("Opción: ")

    # ===== CARGAR =====
    if op == "1":
        cliente = input("Cliente: ")
        auto = input("Auto: ")
        trabajo_desc = input("Trabajo realizado: ")

        total = leer_numero("Total cobrado: ")
        repuestos = leer_numero("Repuestos: ")

        ganancia = total - repuestos

        emp = ganancia * porc_empleado
        inv = ganancia * porc_inversion
        gas = ganancia * porc_gastos
        tu = ganancia - (emp + inv + gas)

        trabajos.append({
            "cliente": cliente,
            "auto": auto,
            "trabajo": trabajo_desc,
            "total": total,
            "repuestos": repuestos,
            "ganancia": ganancia,
            "empleado": emp,
            "inversion": inv,
            "gastos": gas,
            "tu": tu
        })

        print("\nTrabajo cargado")
        print("Tu ganancia:", formatear(tu))

    # ===== RESUMEN =====
    elif op == "2":
        total = sum(t["total"] for t in trabajos)
        tu = sum(t["tu"] for t in trabajos)

        print("\n--- RESUMEN ---")
        print("Trabajos:", len(trabajos))
        print("Facturado:", formatear(total))
        print("Tu ganancia:", formatear(tu))

    # ===== CAMBIAR PORCENTAJES =====
    elif op == "3":
        print("\nPorcentajes actuales:")
        print(f"Empleado: {porc_empleado*100}%")
        print(f"Inversión: {porc_inversion*100}%")
        print(f"Gastos: {porc_gastos*100}%")

        porc_empleado = float(input("Nuevo % empleado (ej 0.20): "))
        porc_inversion = float(input("Nuevo % inversión: "))
        porc_gastos = float(input("Nuevo % gastos: "))

        print("Porcentajes actualizados")

    # ===== OBJETIVO =====
    elif op == "4":
        objetivo = leer_numero("¿Cuánto querés ganar?: ")

        promedio = sum(t["tu"] for t in trabajos) / len(trabajos) if trabajos else 0

        if promedio == 0:
            print("No hay datos")
        else:
            necesarios = objetivo / promedio
            print(f"Necesitás aprox {int(necesarios)+1} trabajos")

    # ===== EXPORTAR A EXCEL PRO =====
    elif op == "5":

        escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
        carpeta = os.path.join(escritorio, "Taller")
        os.makedirs(carpeta, exist_ok=True)

        fecha = datetime.now().strftime("%Y-%m-%d_%H-%M")
        archivo = os.path.join(carpeta, f"taller_{fecha}.xlsx")

        wb = Workbook()
        ws = wb.active
        ws.title = "Resumen Taller"

        # ENCABEZADOS
        encabezados = [
            "Cliente", "Auto", "Trabajo",
            "Total", "Repuestos", "Ganancia",
            "Empleado", "Inversión", "Gastos", "Tu ganancia"
        ]
        ws.append(encabezados)

        # DATOS
        for t in trabajos:
            ws.append([
                t["cliente"],
                t["auto"],
                t["trabajo"],
                t["total"],
                t["repuestos"],
                t["ganancia"],
                t["empleado"],
                t["inversion"],
                t["gastos"],
                t["tu"]
            ])

        # TOTALES
        fila_total = len(trabajos) + 2
        ws[f"A{fila_total}"] = "TOTALES"

        ws[f"D{fila_total}"] = f"=SUM(D2:D{fila_total-1})"
        ws[f"F{fila_total}"] = f"=SUM(F2:F{fila_total-1})"
        ws[f"G{fila_total}"] = f"=SUM(G2:G{fila_total-1})"
        ws[f"J{fila_total}"] = f"=SUM(J2:J{fila_total-1})"

        wb.save(archivo)

        print("\nExcel PRO generado en:")
        print(archivo)

    # ===== SALIR =====
    elif op == "6":
        break

    else:
        print("Opción inválida")